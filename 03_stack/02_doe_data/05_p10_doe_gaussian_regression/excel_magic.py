# -*- coding: utf-8 -*-
"""
Created on Fri Jul 26 12:47:08 2024

@author: wenzel.gassner
"""

import openpyxl
from openpyxl.styles import Font, Border, Side
from openpyxl import Workbook
import os

def initialize_excel_file(file_path="test.xlsx"):
    """
    Create excel file with bold headder if not existing.

    Parameters
    ----------
    file_path : TYPE, optional
        The default is "test.xlsx".
        TODO: think about a good bath later....
    Returns
    -------
    None.

    """
    # Ensure the directory exists
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Optimization Results"
        # Define header
        header = [
            "Power Constraint (kW)", "Cell Count", "Flight Level (100ft)",
            "Feature Name", "Optimized Value", "Lower Bound", "Upper Bound",
            "Hydrogen Consumption (g/s)", "Cell Voltage (V)", "System Power (kW)",
            "Compressor Power (kW)", "Stack Power (kW)"
        ]
        ws.append(header)

        # Style the header
        header_font = Font(bold=True)
        thick_border = Border(
            left=Side(style='thick'), right=Side(style='thick'), 
            top=Side(style='thick'), bottom=Side(style='thick')
        )
        for cell in ws[1]:
            cell.font = header_font
            cell.border = thick_border
            

        wb.save(file_path)
        print(f"Created new Excel file with headers at {file_path}")

def write_to_excel(file_path, data):
    """
    Write data to existing excel file 

    Parameters
    ----------
    file_path : TYPE
        DESCRIPTION.
    data : TYPE
        Should be a list aff all the opt. outputs and inputs:
            data_row = [
                power_constraint_kW, specified_cell_count, flight_level_100ft,
                name, value, bound[0], bound[1],
                hydrogen_mass_flow_g_s, cell_voltage, system_power_kW,
                compressor_power_kW, stack_power_kW
            ]

    Returns
    -------
    None.

    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    ws.append(data)
    wb.save(file_path)

def save_results_to_excel(file_path, feature_names, optimal_input, bounds, hydrogen_mass_flow_g_s, cell_voltage, system_power_kW, compressor_power_kW, stack_power_kW, power_constraint_kW, specified_cell_count, flight_level_100ft):
    """
    Eats all output and input data and appends them to an excel file.

    Parameters
    ----------
    file_path : TYPE
        DESCRIPTION.
    feature_names : TYPE
        DESCRIPTION.
    optimal_input : TYPE
        DESCRIPTION.
    bounds : TYPE
        DESCRIPTION.
    hydrogen_mass_flow_g_s : TYPE
        DESCRIPTION.
    cell_voltage : TYPE
        DESCRIPTION.
    system_power_kW : TYPE
        DESCRIPTION.
    compressor_power_kW : TYPE
        DESCRIPTION.
    stack_power_kW : TYPE
        DESCRIPTION.
    power_constraint_kW : TYPE
        DESCRIPTION.
    specified_cell_count : TYPE
        DESCRIPTION.
    flight_level_100ft : TYPE
        DESCRIPTION.

    Returns
    -------
    None.

    """
    # Iterate over the optimized input variables and their bounds to save them
    for name, value, bound in zip(feature_names, optimal_input, bounds):
        data_row = [
            power_constraint_kW, specified_cell_count, flight_level_100ft,
            name, value, bound[0], bound[1],
            hydrogen_mass_flow_g_s, cell_voltage, system_power_kW,
            compressor_power_kW, stack_power_kW
        ]
        write_to_excel(file_path, data_row)

    # print(f"\nOptimization Stack Power Constraint: {power_constraint_kW:.0f} kW")
    # print(f"Specified Flight Level: {flight_level_100ft} (100x ft)")
    # print(f"Specified Cell Count: {specified_cell_count}")
    # print(f"Calculated and saved data for {power_constraint_kW} kW with {specified_cell_count} cells at flight level {flight_level_100ft}.")